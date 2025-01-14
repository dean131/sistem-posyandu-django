from django.http import JsonResponse
from .models import ChildMeasurement, AnthropometricStandard
from django.views.generic import View
from openpyxl import load_workbook
from django.contrib import messages
from django.shortcuts import redirect, render


class ChildMeasurementListView(View):
    def get(self, request, activity_id, *args, **kwargs):
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            measurements = ChildMeasurement.objects.filter(
                posyandu_activity_id=activity_id
            )
            search_value = request.GET.get("search[value]", "").strip()
            start = int(request.GET.get("start", 0))
            length = int(request.GET.get("length", 10))

            # Apply search filter
            if search_value:
                measurements = measurements.filter(
                    child__full_name__icontains=search_value
                )

            total_records = measurements.count()
            measurements = measurements[start : start + length]

            # Format data for DataTable
            data = [
                {
                    "id": measurement.id,
                    "child": {"full_name": measurement.child.full_name},
                    "weight": measurement.weight,
                    "height": measurement.height,
                    "head_circumference": measurement.head_circumference,
                    "age_in_month": measurement.age_in_month,
                    "measurement_method": measurement.get_measurement_method_display(),
                }
                for measurement in measurements
            ]

            return JsonResponse(
                {
                    "draw": int(request.GET.get("draw", 0)),
                    "recordsTotal": total_records,
                    "recordsFiltered": total_records,
                    "data": data,
                }
            )

        return JsonResponse({"error": "Invalid request"}, status=400)


class AnthropometricStandardView(View):
    def get(self, request, *args, **kwargs):
        return render(request, "child_measurement/anthropometric_standard.html")

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get("excel_file")
        print(excel_file)
        measurement_type = request.POST.get("measurement_type")
        gender = request.POST.get("gender")

        # memastikan file tidak kosong
        if not excel_file:
            messages.error(request, "File Excel dibutuhkan")
            return redirect("anthropometric_standard")

        workbook = load_workbook(excel_file)
        for row in workbook.worksheets[0].iter_rows(values_only=True, min_row=1):
            obj, created = AnthropometricStandard.objects.update_or_create(
                index=row[0],
                sd_minus_3=row[1],
                sd_minus_2=row[2],
                sd_minus_1=row[3],
                median=row[4],
                sd_plus_1=row[5],
                sd_plus_2=row[6],
                sd_plus_3=row[7],
                measurement_type=measurement_type,
                gender=gender,
            )
            if created:
                print("DATA BARU")
            else:
                print("DATA SUDAH ADA")
        messages.success(request, "Data berhasil diupload")
        return redirect("anthropometric_standard")
