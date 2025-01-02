from openpyxl import load_workbook

from django.shortcuts import render
from django.views import View
from django.shortcuts import redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db.models import Prefetch

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator

from account.models import (
    Puskesmas,
    Parent,
    Midwife,
    Cadre,
)


from village.models import Village
from child.models import Child
from posyandu.models import Posyandu
from posyandu_activity.models import PosyanduActivity
from child_measurement.models import AnthropometricStandard


class LoginView(View):
    def get(self, request, *args, **kwargs):
        next_page = request.GET.get("next", "/")
        if request.user.is_authenticated:
            # Redirect to previous page or home page
            return redirect(next_page if next_page != "/login/" else "/")
        return render(request, "adminapp/auth_login.html")

    def post(self, request, *args, **kwargs):
        email = request.POST.get("email")
        password = request.POST.get("password")

        user = Puskesmas.objects.filter(email=email).first()
        if user is not None and user.role == "PUSKESMAS":
            is_password_valid = user.check_password(password)
            if is_password_valid:
                login(request, user)
                return redirect("dashboard")

        messages.error(request, "Email atau password salah")
        return redirect("login")


class LogoutView(View):
    def get(self, request, *args, **kwargs):
        logout(request)
        return redirect("login")


@method_decorator(login_required(login_url="login"), name="dispatch")
class DashboardView(View):
    def get(self, request, *args, **kwargs):
        # Use related objects efficiently to minimize query counts
        user_villages = request.user.village_set.prefetch_related(
            Prefetch(
                "posyandu_set",
                queryset=Posyandu.objects.only("id"),
            )
        )

        # Use aggregation to count related data more efficiently
        posyandu_ids = Posyandu.objects.filter(village__in=user_villages).values_list(
            "id", flat=True
        )
        village_ids = user_villages.values_list("id", flat=True)

        # Aggregate counts with fewer queries
        children_count = (
            Child.objects.filter(parent__posyandus__id__in=posyandu_ids)
            .distinct()
            .count()
        )
        parents_count = (
            Parent.objects.filter(posyandus__id__in=posyandu_ids).distinct().count()
        )
        midwives_count = (
            Midwife.objects.filter(villages__id__in=village_ids).distinct().count()
        )
        cadres_count = (
            Cadre.objects.filter(posyandus__id__in=posyandu_ids).distinct().count()
        )

        # Pass data to the context
        context = {
            "children": children_count,
            "parents": parents_count,
            "midwives": midwives_count,
            "cadres": cadres_count,
        }
        return render(request, "adminapp/dashboard.html", context)


class PosyanduActivitiesView(View):
    def get(self, request, *args, **kwargs):
        village_id = request.GET.get("village_id", "")
        posyandu_id = request.GET.get("posyandu_id", "")

        if village_id:
            village = Village.objects.filter(id=village_id).first()
            posyandu_activities = PosyanduActivity.objects.filter(
                posyandu__village__puskesmas=request.user,
                posyandu__village__name=village,
            ).order_by("-created_at")

        if posyandu_id:
            posyandu = Posyandu.objects.filter(id=posyandu_id).first()
            posyandu_activities = posyandu_activities.filter(
                posyandu=posyandu
            ).order_by("-created_at")

        if not village_id and not posyandu_id:
            posyandu_activities = PosyanduActivity.objects.filter(
                posyandu__village__puskesmas=request.user
            ).order_by("-created_at")

        context = {"posyandu_activities": posyandu_activities}
        return render(request, "adminapp/posyandu_activities.html", context)


from django.views.generic.list import ListView
from django.db.models import Q


class VillageListView(ListView):
    model = Village
    template_name = "adminapp/village_list.html"
    context_object_name = "villages"
    paginate_by = 1  # Add pagination to improve usability

    def get_queryset(self):
        query = self.request.GET.get("q", "").strip()
        puskesmas = self.request.user

        # Filter villages by the user's Puskesmas and optional search query
        queryset = Village.objects.filter(puskesmas=puskesmas).order_by("-created_at")
        if query:
            queryset = queryset.filter(name__icontains=query)

        return queryset

    def get_context_data(self, **kwargs):
        # Add the search query back to the context for template use
        context = super().get_context_data(**kwargs)
        context["query"] = self.request.GET.get("q", "").strip()
        return context

    def post(self, request, *args, **kwargs):
        action = request.POST.get("form-action")
        match action:
            case "add":
                self.add_village(request)
            case "edit":
                self.edit_village(request)
            case "delete":
                self.delete_village(request)
        return redirect("village_list")

    def add_village(self, request):
        name = request.POST.get("village-name")
        Village.objects.create(puskesmas=request.user, name=name)
        messages.success(request, "Desa berhasil ditambahkan")
        return redirect("village_list")

    def delete_village(self, request):
        village_id = request.POST.get("pk")
        village = Village.objects.filter(id=village_id).first()
        if village is not None:
            village.delete()
            messages.success(request, "Desa berhasil dihapus")
        else:
            messages.error(request, "Desa tidak ditemukan")
        return redirect("village_list")

    def edit_village(self, request):
        village_id = request.POST.get("pk")
        name = request.POST.get("village-name")
        village = Village.objects.filter(id=village_id).first()
        if village is not None:
            village.name = name
            village.save()
            messages.success(request, "Desa berhasil diubah")
        else:
            messages.error(request, "Desa tidak ditemukan")
        return redirect("village_list")


class VillageInfoView(View):
    def get(self, request, *args, **kwargs):
        pk = kwargs.get("pk")
        village = Village.objects.filter(id=pk).first()
        children = Child.objects.filter(parent__address__village=village.name)
        parents = Parent.objects.filter(address__village=village.name)
        posyandus = Posyandu.objects.filter(village=village)
        midwives = Midwife.objects.filter(midwifeassignment__village=village)
        context = {
            "village": village,
            "children": children,
            "parents": parents,
            "posyandus": posyandus,
            "midwives": midwives,
        }
        return render(request, "adminapp/village_info.html", context)


class PosyanduView(View):
    def get(self, request, *args, **kwargs):
        query = request.GET.get("q", "")

        if query:
            posyandus = Posyandu.objects.filter(
                village__puskesmas=request.user, name__icontains=query
            ).order_by("-created_at")
        else:
            posyandus = Posyandu.objects.filter(
                village__puskesmas=request.user
            ).order_by("-created_at")

        context = {"posyandus": posyandus}
        return render(request, "adminapp/posyandu_page.html", context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("form-action")
        match action:
            case "add":
                self.add_posyandu(request)
            case "edit":
                self.edit_posyandu(request)
            case "delete":
                self.delete_posyandu(request)
        return redirect("posyandu")

    def add_posyandu(self, request):
        name = request.POST.get("posyandu-name")
        address = request.POST.get("posyandu-address")
        village_id = request.POST.get("village-id")
        village = Village.objects.filter(id=village_id).first()
        if village is not None:
            Posyandu.objects.create(name=name, address=address, village=village)
            messages.success(request, "Posyandu berhasil ditambahkan")
        else:
            messages.error(request, "Desa tidak ditemukan")
        return redirect("posyandu")

    def edit_posyandu(self, request):
        posyandu_id = request.POST.get("pk")
        name = request.POST.get("posyandu-name")
        address = request.POST.get("posyandu-address")
        village_id = request.POST.get("village-id")
        posyandu = Posyandu.objects.filter(id=posyandu_id).first()
        village = Village.objects.filter(id=village_id).first()
        if posyandu is not None:
            posyandu.name = name
            posyandu.address = address
            posyandu.village = village
            posyandu.save()
            messages.success(request, "Posyandu berhasil diubah")
        else:
            messages.error(request, "Posyandu tidak ditemukan")

    def delete_posyandu(self, request):
        posyandu_id = request.POST.get("pk")
        posyandu = Posyandu.objects.filter(id=posyandu_id).first()
        if posyandu is not None:
            posyandu.delete()
            messages.success(request, "Posyandu berhasil dihapus")
        else:
            messages.error(request, "Posyandu tidak ditemukan")


class MidwifeView(View):
    def get(self, request, *args, **kwargs):
        query = request.GET.get("q", "")

        if query:
            midwives = Midwife.objects.filter(full_name__icontains=query).order_by(
                "-created_at"
            )
        else:
            midwives = Midwife.objects.filter().order_by("-created_at")

        context = {"midwives": midwives}
        return render(request, "adminapp/midwife.html", context)


class MidwifeInfoView(View):
    def get(self, request, *args, **kwargs):
        pk = kwargs.get("pk")
        midwife = Midwife.objects.filter(id=pk).first()
        context = {"midwife": midwife}
        return render(request, "adminapp/midwife_info.html", context)


class MidwifeAssignmentView(View):
    def post(self, request, *args, **kwargs):
        action = request.POST.get("form-action")
        midwife_id = request.POST.get("midwife-id")
        match action:
            case "add":
                self.add_midwifeassignment(request)
            case "delete":
                self.delete_midwifeassignment(request)
        return redirect("midwife_info", pk=midwife_id)

    def add_midwifeassignment(self, request):
        midwife_id = request.POST.get("midwife-id")
        village_id = request.POST.get("village-id")

        is_exist = MidwifeAssignment.objects.filter(
            midwife__id=midwife_id, village__id=village_id
        ).exists()
        if is_exist:
            messages.success(request, "Bidan sudah ditambahkan")
            return

        midwife = Midwife.objects.filter(id=midwife_id).first()
        village = Village.objects.filter(id=village_id).first()
        if midwife is not None and village is not None:
            MidwifeAssignment.objects.create(midwife=midwife, village=village)
            messages.success(request, "Bidan berhasil ditambahkan")
        else:
            messages.error(request, "Bidan atau desa tidak ditemukan")

    def delete_midwifeassignment(self, request):
        midwifeassignment_id = request.POST.get("pk")
        midwifeassignment = MidwifeAssignment.objects.filter(
            id=midwifeassignment_id
        ).first()
        if midwifeassignment is not None:
            midwifeassignment.delete()
            messages.success(request, "Penugasan berhasil dihapus")
        else:
            messages.error(request, "Penugasan tidak ditemukan")


class CadreView(View):
    def get(self, request, *args, **kwargs):
        query = request.GET.get("q", "")

        if query:
            cadres = Cadre.objects.filter(full_name__icontains=query).order_by(
                "-created_at"
            )
        else:
            cadres = Cadre.objects.filter().order_by("-created_at")

        context = {"cadres": cadres}
        return render(request, "adminapp/cadres.html", context)


class CadreAssignmentView(View):
    def post(self, request, *args, **kwargs):
        action = request.POST.get("form-action")
        cadre_id = request.POST.get("cadre-id")
        match action:
            case "add":
                self.add_cadreassignment(request)
            case "delete":
                self.delete_cadreassignment(request)
        return redirect("cadre_info", pk=cadre_id)

    def add_cadreassignment(self, request):
        cadre_id = request.POST.get("cadre-id")
        posyandu_id = request.POST.get("posyandu-id")

        is_exist = CadreAssignment.objects.filter(
            cadre__id=cadre_id, posyandu__id=posyandu_id
        ).exists()

        if is_exist:
            messages.success(request, "Kader sudah ditambahkan")
            return

        cadre = Cadre.objects.filter(id=cadre_id).first()
        posyandu = Posyandu.objects.filter(id=posyandu_id).first()
        if cadre is not None and posyandu is not None:
            CadreAssignment.objects.create(cadre=cadre, posyandu=posyandu)
            messages.success(request, "Kader berhasil ditambahkan")
        else:
            messages.error(request, "Kader atau posyandu tidak ditemukan")

    def delete_cadreassignment(self, request):
        cadreassignment_id = request.POST.get("pk")
        cadreassignment = CadreAssignment.objects.filter(id=cadreassignment_id).first()
        if cadreassignment is not None:
            cadreassignment.delete()
            messages.success(request, "Penugasan berhasil dihapus")
        else:
            messages.error(request, "Penugasan tidak ditemukan")


# class ChildView(View):
#     def get(self, request, *args, **kwargs):
#         query = request.GET.get("q", "")

#         posyandus = []
#         for village in request.user.village_set.all():
#             for posyandu in village.posyandu_set.all():
#                 posyandus.append(posyandu)

#         if query:
#             children = Child.objects.filter(
#                 parent__parentposyandu__posyandu__in=posyandus,
#                 full_name__icontains=query,
#             ).order_by("-created_at")
#         else:
#             children = Child.objects.filter(
#                 parent__parentposyandu__posyandu__in=posyandus
#             ).order_by("-created_at")

#         context = {"children": children}
#         return render(request, "adminapp/children.html", context)

#     def post(self, request, *args, **kwargs):
#         action = request.POST.get("form-action")
#         match action:
#             case "add":
#                 self.add_child(request)
#             case "edit":
#                 self.edit_child(request)
#             case "delete":
#                 self.delete_child(request)
#         return redirect("children")


class ParentView(View):
    def get(self, request, *args, **kwargs):
        query = request.GET.get("q", "")

        if query:
            parents = Parent.objects.filter(
                address__subdistrict=request.user.address.subdistrict,
                full_name__icontains=query,
            ).order_by("-created_at")
        else:
            parents = Parent.objects.filter(
                address__subdistrict=request.user.address.subdistrict
            ).order_by("-created_at")

        context = {"parents": parents}
        return render(request, "adminapp/parents.html", context)


class AnthropometricStandardView(View):
    def get(self, request, *args, **kwargs):
        return render(request, "adminapp/anthropometric_standard.html")

    def post(self, request, *args, **kwargs):
        excel_files = request.FILES.getlist("files")
        # memastikan file tidak kosong
        if not excel_files:
            messages.error(request, "File Excel dibutuhkan")
            return redirect("anthropometric_standard")

        array_of_measurement_type = [
            "f_weight_for_age",
            "f_weight_for_length",
            "f_weight_for_height",
            "f_length_for_age",
            "f_height_for_age",
            "f_bmi_for_age_0_24",
            "f_bmi_for_age_24_60",
            "m_weight_for_age",
            "m_weight_for_length",
            "m_weight_for_height",
            "m_length_for_age",
            "m_height_for_age",
            "m_bmi_for_age_0_24",
            "m_bmi_for_age_24_60",
        ]

        # cek measurement type apakah terdaftar
        for excel_file in excel_files:
            workbook = load_workbook(excel_file)
            # cek menggunakan nama file
            file_name = excel_file.name.split(".")[0]
            # jika nama file tidak terdaftar maka akan muncul pesan error
            if file_name not in array_of_measurement_type:
                messages.error(request, f"Measurement type {file_name} tidak terdaftar")
                return redirect("anthropometric_standard")

            for row in workbook.worksheets[0].iter_rows(values_only=True, min_row=2):
                obj = AnthropometricStandard.objects.filter(
                    index=row[0], measurement_type=file_name
                )
                if obj.exists():
                    obj.update(
                        index=row[0],
                        sd_minus_3=row[1],
                        sd_minus_2=row[2],
                        sd_minus_1=row[3],
                        median=row[4],
                        sd_plus_1=row[5],
                        sd_plus_2=row[6],
                        sd_plus_3=row[7],
                        measurement_type=file_name,
                    )
                else:
                    AnthropometricStandard.objects.create(
                        index=row[0],
                        sd_minus_3=row[1],
                        sd_minus_2=row[2],
                        sd_minus_1=row[3],
                        median=row[4],
                        sd_plus_1=row[5],
                        sd_plus_2=row[6],
                        sd_plus_3=row[7],
                        measurement_type=file_name,
                    )
        messages.success(request, "Data berhasil diupload")
        return redirect("anthropometric_standard")


# class LengthForAgeBoysView(View):
#     def get(self, request, *args, **kwargs):
#         return render(request, "adminapp/length_for_age_boys.html")

#     def post(self, request, *args, **kwargs):
#         excel_file = request.FILES.get("file")
#         if excel_file is not None:
#             excel_file = load_workbook(excel_file)
#             for row in excel_file.worksheets[0].iter_rows(values_only=True):
#                 obj = LengthForAgeBoys.objects.filter(age_months=row[0])
#                 if obj.exists():
#                     obj.update(
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#                 else:
#                     LengthForAgeBoys.objects.create(
#                         age_months=row[0],
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#             messages.success(request, "Data berhasil diupload")
#         else:
#             messages.error(request, "File tidak ditemukan")
#         return redirect("length_for_age_boys")


# class LengthForAgeGirlsView(View):
#     def get(self, request, *args, **kwargs):
#         return render(request, "adminapp/length_for_age_girls.html")

#     def post(self, request, *args, **kwargs):
#         excel_file = request.FILES.get("file")
#         if excel_file is not None:
#             excel_file = load_workbook(excel_file)
#             for row in excel_file.worksheets[0].iter_rows(values_only=True):
#                 obj = LengthForAgeGirls.objects.filter(age_months=row[0])
#                 if obj.exists():
#                     obj.update(
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#                 else:
#                     LengthForAgeGirls.objects.create(
#                         age_months=row[0],
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#             messages.success(request, "Data berhasil diupload")
#         else:
#             messages.error(request, "File tidak ditemukan")
#         return redirect("length_for_age_girls")


# class HeightForAgeBoysView(View):
#     def get(self, request, *args, **kwargs):
#         return render(request, "adminapp/height_for_age_boys.html")

#     def post(self, request, *args, **kwargs):
#         excel_file = request.FILES.get("file")
#         if excel_file is not None:
#             excel_file = load_workbook(excel_file)
#             for row in excel_file.worksheets[0].iter_rows(values_only=True):
#                 obj = HeightForAgeBoys.objects.filter(age_months=row[0])
#                 if obj.exists():
#                     obj.update(
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#                 else:
#                     HeightForAgeBoys.objects.create(
#                         age_months=row[0],
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#             messages.success(request, "Data berhasil diupload")
#         else:
#             messages.error(request, "File tidak ditemukan")
#         return redirect("height_for_age_boys")


# class HeightForAgeGirlsView(View):
#     def get(self, request, *args, **kwargs):
#         return render(request, "adminapp/height_for_age_girls.html")

#     def post(self, request, *args, **kwargs):
#         excel_file = request.FILES.get("file")
#         if excel_file is not None:
#             excel_file = load_workbook(excel_file)
#             for row in excel_file.worksheets[0].iter_rows(values_only=True):
#                 obj = HeightForAgeGirls.objects.filter(age_months=row[0])
#                 if obj.exists():
#                     obj.update(
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#                 else:
#                     HeightForAgeGirls.objects.create(
#                         age_months=row[0],
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#             messages.success(request, "Data berhasil diupload")
#         else:
#             messages.error(request, "File tidak ditemukan")
#         return redirect("height_for_age_girls")


# class WeightForAgeBoysView(View):
#     def get(self, request, *args, **kwargs):
#         return render(request, "adminapp/weight_for_age_boys.html")

#     def post(self, request, *args, **kwargs):
#         excel_file = request.FILES.get("file")
#         if excel_file is not None:
#             excel_file = load_workbook(excel_file)
#             for row in excel_file.worksheets[0].iter_rows(values_only=True):
#                 obj = WeightForAgeBoys.objects.filter(age_months=row[0])
#                 if obj.exists():
#                     obj.update(
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#                 else:
#                     WeightForAgeBoys.objects.create(
#                         age_months=row[0],
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#             messages.success(request, "Data berhasil diupload")
#         else:
#             messages.error(request, "File tidak ditemukan")
#         return redirect("weight_for_age_boys")


# class WeightForAgeGirlsView(View):
#     def get(self, request, *args, **kwargs):
#         return render(request, "adminapp/weight_for_age_girls.html")

#     def post(self, request, *args, **kwargs):
#         excel_file = request.FILES.get("file")
#         if excel_file is not None:
#             excel_file = load_workbook(excel_file)
#             for row in excel_file.worksheets[0].iter_rows(values_only=True):
#                 obj = WeightForAgeGirls.objects.filter(age_months=row[0])
#                 if obj.exists():
#                     obj.update(
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#                 else:
#                     WeightForAgeGirls.objects.create(
#                         age_months=row[0],
#                         sd_minus_3=row[1],
#                         sd_minus_2=row[2],
#                         sd_minus_1=row[3],
#                         median=row[4],
#                         sd_plus_1=row[5],
#                         sd_plus_2=row[6],
#                         sd_plus_3=row[7]
#                     )
#             messages.success(request, "Data berhasil diupload")
#         else:
#             messages.error(request, "File tidak ditemukan")
#         return redirect("weight_for_age_girls")
